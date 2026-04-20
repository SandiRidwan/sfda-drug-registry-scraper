"""
╔══════════════════════════════════════════════════════════════════════╗
║   SFDA REGISTERED DRUGS SCRAPER — PRODUCTION READY                 ║
║   API: https://sfda.gov.sa/GetDrugs.php?page=N                     ║
║                                                                      ║
║   Verified dari audit:                                               ║
║     • Total records : 8,756                                          ║
║     • Total pages   : 876 (pageSize=10)                             ║
║     • Method        : GET                                            ║
║     • Auth required : Tidak                                          ║
║     • WAF           : Tidak terdeteksi                               ║
║                                                                      ║
║   Output: sfda_registered_drugs.xlsx                                ║
║   Usage : python sfda_scraper.py                                    ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import requests
import pandas as pd
import time
import random
import json
import sys
import logging
from datetime import datetime
from pathlib import Path

# ── LOGGING ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("sfda_scraper.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ── KONFIGURASI ───────────────────────────────────────────────────────────────

API_URL     = "https://sfda.gov.sa/GetDrugs.php"
OUTPUT_FILE = "sfda_registered_drugs.xlsx"
CHECKPOINT  = "sfda_checkpoint.json"   # resume jika scraping terhenti

# Dari audit: pageSize=10, tapi kita coba naikkan ke 100 untuk efisiensi
# Fallback ke 10 jika server menolak pageSize custom
PAGE_SIZE_REQUEST = 100   # akan dicoba dulu; lihat fungsi probe_page_size()

# Delay antar request (detik)
DELAY_MIN = 0.8
DELAY_MAX = 2.0

# Retry config
MAX_RETRIES   = 5
BACKOFF_BASE  = 3   # detik, dikali 2^attempt saat retry

# Header persis seperti yang direcord di DevTools
HEADERS = {
    "Accept":          "application/json, text/javascript, */*; q=0.01",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Referer":         "https://www.sfda.gov.sa/en/drugs-list",
    "Connection":      "keep-alive",
    "DNT":             "1",
}

# ── MAPPING FIELD (dari audit JSON) ──────────────────────────────────────────
# Semua field yang tersedia di API → nama kolom di Excel
FIELD_MAP = {
    # 5 field utama yang diminta klien
    "tradeName":                  "Drug Name",
    "manufacturerName":           "Manufacturing Company",
    "manufacturerCountry":        "Country of Manufacture",
    "agent":                      "Marketing Company / Saudi Agent",
    "registerNumber":             "Registration Number",

    # Field tambahan yang bisa berguna (keep untuk kelengkapan data)
    "scientificName":             "Scientific Name",
    "domainEN":                   "Drug Domain",
    "drugType":                   "Drug Type",
    "doesageForm":                "Dosage Form",
    "administrationRoute":        "Administration Route",
    "strength":                   "Strength",
    "strengthUnit":               "Strength Unit",
    "packageType":                "Package Type",
    "packageSize":                "Package Size",
    "size":                       "Volume / Size",
    "sizeUnit":                   "Size Unit",
    "price":                      "Price (SAR)",
    "marketingStatus":            "Marketing Status",
    "authorizationStatus":        "Authorization Status",
    "legalStatusEn":              "Legal Status",
    "productControl":             "Product Control",
    "distributionArea":           "Distribution Area",
    "storageConditions":          "Storage Conditions",
    "shelfLife":                  "Shelf Life (months)",
    "registerYear":               "Registration Year",
    "atcCode1":                   "ATC Code 1",
    "atcCode2":                   "ATC Code 2",
    "companyName":                "Company Name",
    "companyCountryEn":           "Company Country",
    "additionalManufacturer":     "Additional Manufacturer",
    "additionalManufacturerCountry": "Additional Manufacturer Country",
}

# Kolom output yang diurutkan (5 utama di depan)
COLUMN_ORDER = [
    "Drug Name",
    "Manufacturing Company",
    "Country of Manufacture",
    "Marketing Company / Saudi Agent",
    "Registration Number",
    "Scientific Name",
    "Drug Domain",
    "Drug Type",
    "Dosage Form",
    "Administration Route",
    "Strength",
    "Strength Unit",
    "Package Type",
    "Package Size",
    "Volume / Size",
    "Size Unit",
    "Price (SAR)",
    "Marketing Status",
    "Authorization Status",
    "Legal Status",
    "Product Control",
    "Distribution Area",
    "Storage Conditions",
    "Shelf Life (months)",
    "Registration Year",
    "ATC Code 1",
    "ATC Code 2",
    "Company Name",
    "Company Country",
    "Additional Manufacturer",
    "Additional Manufacturer Country",
]

# ══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def fetch_page(session: requests.Session, page: int, page_size: int) -> dict | None:
    """
    Fetch satu halaman dari API dengan retry + exponential backoff.
    Returns dict response atau None jika semua retry gagal.
    """
    params = {"page": page, "pageSize": page_size}

    for attempt in range(MAX_RETRIES):
        try:
            resp = session.get(
                API_URL,
                params=params,
                headers=HEADERS,
                timeout=30,
            )

            if resp.status_code == 200:
                data = resp.json()
                # Validasi response punya struktur yang benar
                if "results" in data and isinstance(data["results"], list):
                    return data
                else:
                    log.warning(f"Page {page}: response JSON tidak punya key 'results'. Keys: {list(data.keys())}")
                    return None

            elif resp.status_code == 429:
                wait = BACKOFF_BASE * (2 ** attempt)
                log.warning(f"Page {page}: Rate limited (429). Tunggu {wait}s...")
                time.sleep(wait)

            elif resp.status_code in (500, 502, 503, 504):
                wait = BACKOFF_BASE * (2 ** attempt)
                log.warning(f"Page {page}: Server error {resp.status_code}. Retry {attempt+1}/{MAX_RETRIES} dalam {wait}s...")
                time.sleep(wait)

            else:
                log.error(f"Page {page}: HTTP {resp.status_code}. Berhenti retry.")
                return None

        except requests.exceptions.Timeout:
            wait = BACKOFF_BASE * (2 ** attempt)
            log.warning(f"Page {page}: Timeout. Retry {attempt+1}/{MAX_RETRIES} dalam {wait}s...")
            time.sleep(wait)

        except requests.exceptions.ConnectionError as e:
            wait = BACKOFF_BASE * (2 ** attempt)
            log.warning(f"Page {page}: Connection error ({e}). Retry {attempt+1}/{MAX_RETRIES} dalam {wait}s...")
            time.sleep(wait)

        except json.JSONDecodeError:
            log.error(f"Page {page}: Response bukan JSON valid. Skip.")
            return None

        except Exception as e:
            log.error(f"Page {page}: Unexpected error: {e}")
            return None

    log.error(f"Page {page}: Semua {MAX_RETRIES} retry gagal. Skip halaman ini.")
    return None


def probe_page_size(session: requests.Session) -> tuple[int, int, int]:
    """
    Fetch halaman 1 untuk mendapatkan metadata: rowCount, pageCount, pageSize.
    Coba pageSize=100 dulu. Fallback ke default (10) jika gagal.
    Returns: (total_records, actual_page_count, actual_page_size)
    """
    log.info(f"Probing API dengan pageSize={PAGE_SIZE_REQUEST}...")
    data = fetch_page(session, page=1, page_size=PAGE_SIZE_REQUEST)

    if not data:
        log.error("Gagal probe API. Cek koneksi dan URL.")
        sys.exit(1)

    row_count  = data.get("rowCount",  0)
    page_count = data.get("pageCount", 0)
    page_size  = data.get("pageSize",  10)

    if row_count == 0:
        log.warning("rowCount=0 dari API. Kemungkinan pageSize custom tidak didukung. Fallback ke pageSize=10.")
        data = fetch_page(session, page=1, page_size=10)
        row_count  = data.get("rowCount",  8756)
        page_count = data.get("pageCount", 876)
        page_size  = 10

    log.info(f"API confirmed: {row_count:,} records | {page_count} pages | pageSize={page_size}")
    return row_count, page_count, page_size


def detect_arabic(value) -> bool:
    """True jika string mengandung karakter Arab (Unicode U+0600–U+06FF)."""
    if not isinstance(value, str):
        return False
    return any('\u0600' <= c <= '\u06FF' for c in value)


def parse_record(item: dict) -> dict:
    """
    Ekstrak semua field dari satu drug record API.
    Handle None dengan string kosong. Flag field Arab.
    """
    parsed = {}
    for api_key, col_name in FIELD_MAP.items():
        val = item.get(api_key)

        if val is None:
            parsed[col_name] = ""
        elif isinstance(val, str):
            val = val.strip()
            if detect_arabic(val):
                parsed[col_name] = f"[AR_FLAG] {val}"
            else:
                parsed[col_name] = val
        else:
            parsed[col_name] = val

    return parsed


def save_checkpoint(page: int, records: list):
    """Simpan progress ke file JSON agar bisa resume jika terhenti."""
    cp = {
        "last_page":     page,
        "records_count": len(records),
        "timestamp":     datetime.now().isoformat(),
    }
    with open(CHECKPOINT, "w") as f:
        json.dump(cp, f)


def load_checkpoint() -> int:
    """Load checkpoint. Return halaman terakhir yang berhasil, atau 0."""
    if Path(CHECKPOINT).exists():
        try:
            with open(CHECKPOINT) as f:
                cp = json.load(f)
            last = cp.get("last_page", 0)
            log.info(f"Checkpoint ditemukan: resume dari halaman {last + 1} "
                     f"({cp.get('records_count', 0)} records sudah tersimpan)")
            return last
        except Exception:
            return 0
    return 0

# ══════════════════════════════════════════════════════════════════════════════
# EXPORT TO EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def export_excel(records: list, filename: str):
    """Export list of dicts ke file Excel yang bersih dan terformat."""
    log.info(f"Exporting {len(records):,} records ke {filename}...")

    df = pd.DataFrame(records)

    # Pastikan kolom urut sesuai COLUMN_ORDER (tambah kolom baru di akhir jika ada)
    existing_ordered = [c for c in COLUMN_ORDER if c in df.columns]
    extra_cols = [c for c in df.columns if c not in COLUMN_ORDER]
    df = df[existing_ordered + extra_cols]

    # Deduplication berdasarkan Registration Number
    before = len(df)
    df = df.drop_duplicates(subset=["Registration Number"], keep="first")
    after  = len(df)
    if before != after:
        log.info(f"Deduplication: {before - after} duplikat dihapus → {after:,} unique records")

    # Drop baris yang 5 field utamanya semua kosong
    key_cols = ["Drug Name", "Manufacturing Company", "Country of Manufacture",
                "Marketing Company / Saudi Agent", "Registration Number"]
    df = df[~(df[key_cols] == "").all(axis=1)]

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="SFDA Registered Drugs", index=False)

        ws = writer.sheets["SFDA Registered Drugs"]

        # Style header row
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill   = PatternFill("solid", fgColor="1F4E79")
        header_font   = Font(color="FFFFFF", bold=True, size=10)
        header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border   = Border(
            bottom=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
        )

        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = header_align
            cell.border    = thin_border

        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"   # freeze header row

        # Auto-fit column widths (capped)
        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

        # Alternate row shading untuk readability
        light_fill = PatternFill("solid", fgColor="EFF3FB")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = light_fill

        # Auto-filter on header
        ws.auto_filter.ref = ws.dimensions

        # Add metadata sheet
        meta_ws = writer.book.create_sheet("Metadata")
        meta_ws["A1"] = "SFDA Registered Drugs — Scrape Metadata"
        meta_ws["A1"].font = Font(bold=True, size=12)
        meta_data = [
            ("Source URL",    API_URL),
            ("Scrape Date",   datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Records", len(df)),
            ("Script",        "sfda_scraper.py"),
            ("Fields Captured", ", ".join(key_cols)),
            ("AR_FLAG Note",  "Fields prefixed [AR_FLAG] contain Arabic text — review manually"),
        ]
        for i, (label, value) in enumerate(meta_data, start=3):
            meta_ws[f"A{i}"] = label
            meta_ws[f"B{i}"] = value
            meta_ws[f"A{i}"].font = Font(bold=True)
        meta_ws.column_dimensions["A"].width = 22
        meta_ws.column_dimensions["B"].width = 60

    log.info(f"✓ Excel tersimpan: {filename} ({df.shape[0]:,} rows × {df.shape[1]} cols)")
    return len(df)

# ══════════════════════════════════════════════════════════════════════════════
# MAIN SCRAPER
# ══════════════════════════════════════════════════════════════════════════════

def scrape() -> list:
    """Loop semua halaman API, parse records, return list of dicts."""
    session = requests.Session()
    session.headers.update(HEADERS)

    # Probe API untuk mendapatkan pagination metadata
    total_records, total_pages, page_size = probe_page_size(session)

    # Resume dari checkpoint jika ada
    start_page   = load_checkpoint() + 1
    all_records  = []
    skipped_pages = []

    start_time = time.time()
    log.info(f"Mulai scraping dari halaman {start_page}/{total_pages}...")
    log.info(f"Estimasi selesai: {total_pages * ((DELAY_MIN + DELAY_MAX) / 2) / 60:.1f} menit")
    log.info("-" * 60)

    for page in range(start_page, total_pages + 1):
        data = fetch_page(session, page=page, page_size=page_size)

        if data is None:
            log.warning(f"  Page {page}: SKIP (gagal setelah {MAX_RETRIES} retry)")
            skipped_pages.append(page)
            continue

        results = data.get("results", [])
        page_records = [parse_record(item) for item in results]
        all_records.extend(page_records)

        # Progress log setiap 50 halaman
        if page % 50 == 0 or page == total_pages:
            elapsed   = time.time() - start_time
            progress  = page / total_pages * 100
            remaining = (elapsed / page) * (total_pages - page) if page > 0 else 0
            log.info(
                f"  Page {page:>4}/{total_pages} | "
                f"{progress:5.1f}% | "
                f"{len(all_records):>6,} records | "
                f"~{remaining/60:.1f} mnt tersisa"
            )
            save_checkpoint(page, all_records)

        # Human-like delay antar request
        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    # Retry halaman yang dilewati
    if skipped_pages:
        log.warning(f"\nRetrying {len(skipped_pages)} halaman yang dilewati: {skipped_pages}")
        time.sleep(10)  # cooldown sebelum retry
        for page in skipped_pages:
            data = fetch_page(session, page=page, page_size=page_size)
            if data:
                page_records = [parse_record(item) for item in data.get("results", [])]
                all_records.extend(page_records)
                log.info(f"  Retry page {page}: OK ({len(page_records)} records)")
            else:
                log.error(f"  Retry page {page}: MASIH GAGAL. Data mungkin tidak lengkap.")

    elapsed_total = time.time() - start_time
    log.info(f"\nScraping selesai: {len(all_records):,} records dalam {elapsed_total/60:.1f} menit")
    return all_records


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  SFDA DRUGS LIST SCRAPER")
    print(f"  API: {API_URL}")
    print(f"  Target output: {OUTPUT_FILE}")
    print(f"  Waktu mulai: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    try:
        records = scrape()

        if not records:
            log.error("Tidak ada data yang berhasil diextract. Cek log di atas.")
            sys.exit(1)

        final_count = export_excel(records, OUTPUT_FILE)

        # Hapus checkpoint jika berhasil
        if Path(CHECKPOINT).exists():
            Path(CHECKPOINT).unlink()
            log.info("Checkpoint dihapus (scraping berhasil complete).")

        print()
        print("=" * 60)
        print(f"  ✓ SELESAI: {final_count:,} drug registrations")
        print(f"  ✓ File: {OUTPUT_FILE}")
        print(f"  ✓ Log : sfda_scraper.log")
        print("=" * 60)

    except KeyboardInterrupt:
        log.warning("\nDihentikan oleh user (Ctrl+C). Progress tersimpan di checkpoint.")
        sys.exit(0)
    except Exception as e:
        log.exception(f"Fatal error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
